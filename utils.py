import time
import math
import shutil
import statistics
import scipy.io as scio

from PySide2.QtWidgets import QMessageBox
from openpyxl import load_workbook
from deepforest import CascadeForestClassifier
from libsvm.svmutil import *

from saglrstepvar import *


def get_current_time():
    ct = time.time()
    local_time = time.localtime(ct)
    data_head = time.strftime("%H%M%S", local_time)
    data_secs = (ct - int(ct)) * 1000
    time_stamp = "%s%03d" % (data_head, data_secs)
    return time_stamp


def get_data(data):
    data = data.hex()
    for i in range(4):
        bin_data_tmp = bin(int(data[2 * i:2 * i + 2], 16))[2:]
        if len(bin_data_tmp) < 8:
            loss = 8 - len(bin_data_tmp)
            for j in range(loss):
                bin_data_tmp = '0' + bin_data_tmp
        if i == 0:
            data_out = bin_data_tmp
        else:
            data_out = bin_data_tmp + data_out

    sign = int(data_out[0], 2)
    exp = int(data_out[1:9], 2)
    frac = int(data_out[9:32], 2)

    if exp == 0:
        data_out = ((-1) ** sign) * (2 ** (-126)) * (frac / (2 ** 23))
    else:
        data_out = ((-1) ** sign) * (2 ** (exp - 127)) * (1 + (frac / (2 ** 23)))

    return data_out


def label_preprocess(path, subject, num_train):
    num_gesture = 5
    if 'stand2' in subject:
        num_gesture += 1
    num_action = num_gesture * num_train
    trig_dyn = np.zeros((num_gesture * num_train, 3))
    Trigger = []
    trigger_data = []
    filename = path + subject + '/' + subject + '_Trigger.txt'
    with open(filename) as f:
        file = f.readlines()
    for each in file:
        each = each.strip('\n')
        temp = int(each[0:2]) * 3600 * 1000 + int(each[2:4]) * 60 * 1000 + int(each[4:6]) * 1000 + int(each[6:])
        Trigger.append(temp)

    filename = path + subject + '_label/' + subject + '_trigger.xlsx'
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]
    rows = ws.rows
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)
    for i in range(2 + 2 * num_action):
        each = str(content[i][1])
        temp = int(each[0:2]) * 3600 * 1000 + int(each[2:4]) * 60 * 1000 + int(each[4:6]) * 1000 + int(each[6:])
        trigger_data.append(temp)

    filename = path + subject + '_label/' + subject + '_label.mat'
    file = scio.loadmat(file_name=filename)
    randIndex = file['randIndex']

    ges_start_time = 3000  # ms
    if 'stand2' in subject:
        ges_start_time += 1000

    for num in range(num_gesture * num_train):
        nearestIdx = np.argmin(abs(trigger_data[2 * num + 1] - np.array(Trigger)))
        a = nearestIdx * 100 + + ges_start_time + 250
        # a = trigger_data[2 * num + 1] - Trigger[0] + 3250
        trig_dyn[num, 0] = a // 100 * 100
        b = a + 1000
        trig_dyn[num, 1] = b // 100 * 100
        trig_dyn[num, 2] = randIndex[0][num % 5]
        # trig_dyn[num, 2] = randIndex[0][int(num / num_train + 1) - 1]

    data = {'trig_dyn': trig_dyn}
    filename = path + subject + '/' + subject + '_label_preprocess.mat'
    scio.savemat(filename, data)


def data_preprocess2(widget, path, subject, num_train):
    num_gesture = 5
    num_action = num_gesture * num_train

    filename = path + subject + '/' + subject + '_EMGData.txt'
    with open(filename) as f:
        file = f.readlines()
    EMGData = np.zeros((len(file), 8))
    i = 0
    for each in file:
        each = each.strip('\n')
        each = each.split()
        for j in range(8):
            EMGData[i, j] = float(each[j])
        i += 1
    data = {'EMGData': EMGData}
    filename = path + subject + '/' + subject + '_emg_raw.mat'
    scio.savemat(filename, data)

    filename = path + subject + '/' + subject + '_GroData.txt'
    with open(filename) as f:
        file = f.readlines()
    GroData = np.zeros((len(file), 3))
    i = 0
    for each in file:
        each = each.strip('\n')
        each = each.split()
        for j in range(3):
            GroData[i, j] = float(each[j])
        i += 1
    data = {'GroData': GroData}
    filename = path + subject + '/' + subject + '_gro_raw.mat'
    scio.savemat(filename, data)

    filename = path + subject + '/' + subject + '_AccData.txt'
    with open(filename) as f:
        file = f.readlines()
    AccData = np.zeros((len(file), 3))
    i = 0
    for each in file:
        each = each.strip('\n')
        each = each.split()
        for j in range(3):
            AccData[i, j] = float(each[j])
        i += 1
    data = {'AccData': AccData}
    filename = path + subject + '/' + subject + '_acc_raw.mat'
    scio.savemat(filename, data)

    filename = path + subject + '/' + subject + '_label_preprocess.mat'
    file = scio.loadmat(filename)
    trig_dyn = file['trig_dyn']

    RealSt = np.zeros((trig_dyn.shape[0], EMGData.shape[1]))
    endIndex = int(trig_dyn[-1, 1])  # 此处优化了切割数据，避免采集结束后被试做的动作对切割造成干扰
    for nb in range(EMGData.shape[1]):
        emgdata = EMGData[:endIndex, nb]
        # emgdataP[:, nb] = emgdata
        [t0, _, _, ga] = saglrstepvar(emgdata, 200, 100, 100, 'multiple', 1)
        gt = ga[t0]
        dis = np.transpose(np.tile(trig_dyn[:, 0], (len(t0), 1))) - np.tile(t0, (len(trig_dyn[:, 0]), 1))
        loc = np.argmin(abs(dis), axis=0)
        uniqLoc = np.unique(loc)
        for i in range(len(uniqLoc)):
            temp1 = np.where(loc == uniqLoc[i])
            temp = temp1[0]
            locS = np.argsort(gt[temp])
            if len(locS) > 1:
                tempT0 = t0[temp[locS[0:2]]]
            else:
                tempT0 = t0[temp[locS]]
            finT0 = np.argmin(abs(tempT0 - np.tile(trig_dyn[i, 0], (1, len(tempT0)))))
            RealSt[uniqLoc[i], nb] = tempT0[finT0]
    idx1 = (RealSt - np.transpose(np.tile(trig_dyn[:, 1], (np.shape(EMGData)[1], 1))) > 0)
    idx2 = (abs(RealSt - np.transpose(np.tile(trig_dyn[:, 0], (np.shape(EMGData)[1], 1)))) > 1000)
    RealSt[idx1] = np.nan
    RealSt[idx2] = np.nan
    RealSt = np.delete(RealSt, trig_dyn[:, 2] == 5, axis=0)

    RealSt_tmp = np.floor(RealSt / 100)
    finRealSt = np.zeros(num_action)
    for nn in range(num_action):
        a = [i for i in RealSt_tmp[nn, :] if not math.isnan(i)]
        if len(a):
            temp = statistics.mode(a)
            finRealSt[nn] = temp * 100
        else:
            print('You may make gestures at wrong time(earlier or later).')
            dlg = QMessageBox(widget)
            dlg.setWindowTitle('Warning')
            dlg.setText('You may make gestures at wrong time(earlier or later).')
            dlg.setIcon(QMessageBox.Warning)
            dlg.exec_()
            break

    finRealSt = finRealSt.astype(int)

    emgdata = [None] * num_gesture * num_train
    grodata = [None] * num_gesture * num_train
    accdata = [None] * num_gesture * num_train
    gesture_be = finRealSt - 100
    data_en = gesture_be + 1000
    data_be = gesture_be - 2500
    # data_en = trig_dyn[:, 1]
    # data_be = data_en - 3500
    for i in range(num_gesture * num_train):
        emgdata[i] = EMGData[int(data_be[i]):int(data_en[i])]
        grodata[i] = GroData[int(data_be[i]):int(data_en[i])]
        accdata[i] = AccData[int(data_be[i]):int(data_en[i])]
    filename = path + subject + '/' + subject + '_emg_preprocessed.mat'
    data = {'emgdata': emgdata}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_gro_preprocessed.mat'
    data = {'grodata': grodata}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_acc_preprocessed.mat'
    data = {'accdata': accdata}
    scio.savemat(filename, data)

    emg_rest = [None] * num_gesture * num_train
    emg_dyn = [None] * num_gesture * num_train
    gro_rest = [None] * num_gesture * num_train
    gro_dyn = [None] * num_gesture * num_train
    acc_rest = [None] * num_gesture * num_train
    acc_dyn = [None] * num_gesture * num_train
    for i in range(num_gesture * num_train):
        rawemg = emgdata[i]
        rawgro = grodata[i]
        rawacc = accdata[i]
        emg_rest[i] = rawemg[0:2000, :]
        emg_dyn[i] = rawemg[2500:3500, :]
        gro_rest[i] = rawgro[0:2000, :]
        gro_dyn[i] = rawgro[2500:3500, :]
        acc_rest[i] = rawacc[0:2000, :]
        acc_dyn[i] = rawacc[2500:3500, :]
    filename = path + subject + '/' + subject + '_emg_rest.mat'
    data = {'emg_rest': emg_rest}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_emg_dyn.mat'
    data = {'emg_dyn': emg_dyn}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_gro_rest.mat'
    data = {'gro_rest': gro_rest}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_gro_dyn.mat'
    data = {'gro_dyn': gro_dyn}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_acc_rest.mat'
    data = {'acc_rest': acc_rest}
    scio.savemat(filename, data)
    filename = path + subject + '/' + subject + '_acc_dyn.mat'
    data = {'acc_dyn': acc_dyn}
    scio.savemat(filename, data)


def snr_train(widget, path, subject, num_train):
    snr_thresh = 10

    filename = path + subject + '/' + subject + '_emg_rest.mat'
    file = scio.loadmat(filename)
    emg_rest = file['emg_rest']
    filename = path + subject + '/' + subject + '_emg_dyn.mat'
    file = scio.loadmat(filename)
    emg_dyn = file['emg_dyn']

    snr_all = np.zeros((num_train*5, 8))
    for j in range(emg_dyn.shape[0]):
        emg_sig = emg_dyn[j]
        emg_noi = emg_rest[j]
        rms_sig = np.sqrt(np.mean(emg_sig ** 2, axis=0))
        rms_noi = np.sqrt(np.mean(emg_noi ** 2, axis=0))
        snr_all[j, :] = 20 * np.log(rms_sig / rms_noi)

    snr_channel_mean = np.mean(snr_all, axis=0)
    snr_mean = np.mean(snr_all)

    idx_snr = snr_channel_mean > snr_thresh
    num_snr_channel = np.sum(idx_snr == 1)

    if num_snr_channel < 5:
        if num_snr_channel == 0:
            dlg = QMessageBox(widget)
            dlg.setWindowTitle('Error')
            dlg.setText('所有通道信噪比低，无法训练模型')
            dlg.setIcon(QMessageBox.Critical)
            dlg.exec_()
        else:
            dlg = QMessageBox(widget)
            dlg.setWindowTitle('Warning')
            dlg.setText('超过半数通道信噪比低')
            dlg.setIcon(QMessageBox.Warning)
            dlg.exec_()

    return snr_channel_mean, snr_mean


def deep_forest_online_train(path, subject, sourceID, num_train, num_source_train, snr_channel_mean):
    # filename = path + subject + '_' + str(datetime.date.today()) + '/' + subject + '_' + str(
    #     datetime.date.today()) + '_PRlog.mat'
    # file = scio.loadmat(file_name=filename)
    # randIndex = file['randIndex']
    #
    forest_mdl_train = CascadeForestClassifier(max_layers=5, n_estimators=2, n_trees=30, random_state=1, n_jobs=-1,
                                               backend="sklearn", use_predictor=True)

    gesture_len = 1
    window_len = 0.5
    step_len = 0.1
    num_prd = int((gesture_len - window_len) / step_len + 1)
    pca_active = 0
    dim = 20
    fs_emg = 1000
    snr_thresh = 10
    num_ges = 5
    num_train_all = num_train+num_source_train
    num_windows = num_prd * num_ges * num_train_all
    #
    # label_dyn = [0] * 15
    # for i in range(5):
    #     label_dyn[3 * i:3 * i + 3] = [randIndex[0, i]] * 3

    # 第二天数据导入
    filename = path + subject + '/' + subject + '_emg_rest.mat'
    file = scio.loadmat(filename)
    emg_rest = file['emg_rest']
    filename = path + subject + '/' + subject + '_emg_dyn.mat'
    file = scio.loadmat(filename)
    emg_dyn = file['emg_dyn']
    filename = path + subject + '/' + subject + '_label_preprocess.mat'
    file = scio.loadmat(filename)
    emg_label = file['trig_dyn'][:, 2]

    # 第一天数据导入
    filename = path + sourceID + '/' + sourceID + '_emg_rest.mat'
    file = scio.loadmat(filename)
    data_rest = file['emg_rest']
    filename = path + sourceID + '/' + sourceID + '_emg_dyn.mat'
    file = scio.loadmat(filename)
    data_dyn = file['emg_dyn']
    filename = path + sourceID + '/' + sourceID + '_label_preprocess.mat'
    file = scio.loadmat(filename)
    data_label = file['trig_dyn'][:, 2]

    data_feed = np.concatenate([data_dyn, emg_dyn], axis=0)
    data_rest_feed = np.concatenate([data_rest, emg_rest], axis=0)
    label_feed = np.concatenate([data_label, emg_label, emg_label], axis=0)
    # for i in range(5):
    #     idx_rel_train_tmp.append(label_dyn.index(i))
    # idx_train = []
    # for i in range(5):
    #     idx_train.append(idx_rel_train_tmp[i])
    #     idx_train.append(idx_rel_train_tmp[i] + 1)
    #     idx_train.append(idx_rel_train_tmp[i] + 2)
    # # label_train_1 = [1] * len(idx_train) + [0] * len(idx_train)
    # # label_train_2 = [1]*3 + [2]*3 + [3]*3 + [4]*3 + [5]*3
    # label_train_tmp_1 = [1] * 90 + [0] * 90
    # label_train_tmp_2 = [0] * 18 + [1] * 18 + [2] * 18 + [3] * 18 + [4] * 18
    emg_train_rel = data_feed
    emg_train_irr = data_rest_feed
    # idx_snr = snr_channel_mean > snr_thresh
    # num_channel_snr = sum(idx_snr == 1)
    label_list=[]
    EMGfeature_train_rel = np.zeros((4 * 8, num_windows))
    for j in range(emg_train_rel.shape[0]):
        emg = emg_train_rel[j]
        # emg = emg[:,idx_snr]
        var = get_var(emg, window_len, step_len, fs_emg)
        vo = get_vo(emg, window_len, step_len, fs_emg)
        ltkeo = get_ltkeo(emg, window_len, step_len, fs_emg)
        asm = get_asm(emg, window_len, step_len, fs_emg)
        for k in range(6):
            EMGfeature_train_rel[:, j * num_prd + k] = np.concatenate((var[k, :], vo[k, :], ltkeo[k, :], asm[k, :]),
                                                                      axis=0)
        label_list.append([label_feed[j]] * 6)
    label_feed = np.hstack(label_list)

    EMGfeature_train_irr = np.zeros((4 * 8, num_windows))
    for j in range(emg_train_irr.shape[0]):
        emg = emg_train_irr[j][-1 - 1000 + 1:]
        # emg = emg[:,idx_snr]
        var = get_var(emg, window_len, step_len, fs_emg)
        vo = get_vo(emg, window_len, step_len, fs_emg)
        ltkeo = get_ltkeo(emg, window_len, step_len, fs_emg)
        asm = get_asm(emg, window_len, step_len, fs_emg)
        for k in range(6):
            EMGfeature_train_irr[:, j * num_prd + k] = np.concatenate((var[k, :], vo[k, :], ltkeo[k, :], asm[k, :]),
                                                                      axis=0)

    EMGfeature_train_1 = np.concatenate((EMGfeature_train_rel, EMGfeature_train_irr), axis=1)
    label_train_tmp_1 = [1] * num_windows + [0] * num_windows
    # # 归一化待补充
    # EMGfeature_train_norm_1, EMGmean_val_1, EMGstd_val_1, EMGWPCA_1 = feature_train_normalize(EMGfeature_train_1,
    #                                                                                           pca_active, dim)
    # EMGfeature_train_norm_2, EMGmean_val_2, EMGstd_val_2, EMGWPCA_2 = feature_train_normalize(EMGfeature_train_rel,
    #                                                                                           pca_active, dim)
    EMGmdl_1 = svm_train(label_train_tmp_1, EMGfeature_train_1.T, '-t 0')
    train_weight = np.concatenate([1 * np.ones(num_source_train*num_ges*num_prd), 30 * np.ones(num_train*num_ges*num_prd)], axis=0)
    forest_mdl_train.fit(EMGfeature_train_rel.T, label_feed, train_weight)
    # EMGmdl_2 = svm_train(label_train_tmp_2, EMGfeature_train_norm_2.T, '-t 0')
    svm_save_model(subject + 'mdl', EMGmdl_1)
    save_path = path + subject + '/' + subject +'forest'
    # pathlib.Path(save_path).mkdir(parents=True, exist_ok=True)
    if os.path.exists(save_path):
        shutil.rmtree(save_path)
        #删除目录
    forest_mdl_train.save(save_path)

    # data = {'EMGmean_val_1': EMGmean_val_1, 'EMGstd_val_1': EMGstd_val_1, 'EMGWPCA_1': EMGWPCA_1,
    #         'EMGmean_val_2': EMGmean_val_2, 'EMGstd_val_2': EMGstd_val_2, 'EMGWPCA_2': EMGWPCA_2,
    #         'pca_active': pca_active, 'dim': dim, 'idx_snr': idx_snr}
    # filename = path + subject + '/' + subject + '_train_data.mat'
    # scio.savemat(filename, data)


def my_var(sig):
    N = len(sig)
    mu = np.mean(sig)
    var_value = (1 / (N - 1)) * np.sum((sig - mu) ** 2)
    return var_value


def my_vo(sig):
    order = 2
    N = len(sig)
    Y = (1 / N) * np.sum(sig ** order)
    vo_value = Y ** (1 / order)
    return vo_value


def my_ltkeo(sig):
    N = len(sig)
    Y = 0
    for j in range(1, N - 1):
        Y += (sig[j] ** 2) - sig[j - 1] * sig[j + 1]
    ltkeo_value = np.log(Y)
    return ltkeo_value


def my_asm(sig):
    K = len(sig)
    Y = 0
    for n in range(0, K):
        if 0.25 * K <= n + 1 <= 0.75 * K:
            exp = 0.5
        else:
            exp = 0.75
        Y += round(sig[n]) ** exp
    asm_value = np.abs(Y / K)
    return asm_value


def get_var(emg, window_len, step_len, fs):
    window_sample = int(window_len * fs)
    step_sample = int(step_len * fs)
    (Nsample, Nchannel) = emg.shape

    var = np.zeros((6, Nchannel))
    fea_idx = -1
    for i in range(0, Nsample - window_sample + 1, step_sample):
        fea_idx += 1
        for j in range(Nchannel):
            emg_window = emg[i:i + window_sample, j]
            var[fea_idx, j] = my_var(emg_window)
    return var


def get_vo(emg, window_len, step_len, fs):
    window_sample = int(window_len * fs)
    step_sample = int(step_len * fs)
    (Nsample, Nchannel) = emg.shape

    vo = np.zeros((6, Nchannel))
    fea_idx = -1
    for i in range(0, Nsample - window_sample + 1, step_sample):
        fea_idx += 1
        for j in range(Nchannel):
            emg_window = emg[i:i + window_sample, j]
            vo[fea_idx, j] = my_vo(emg_window)
    return vo


def get_ltkeo(emg, window_len, step_len, fs):
    window_sample = int(window_len * fs)
    step_sample = int(step_len * fs)
    (Nsample, Nchannel) = emg.shape

    ltkeo = np.zeros((6, Nchannel))
    fea_idx = -1
    for i in range(0, Nsample - window_sample + 1, step_sample):
        fea_idx += 1
        for j in range(Nchannel):
            emg_window = emg[i:i + window_sample, j]
            ltkeo[fea_idx, j] = my_ltkeo(emg_window)
    return ltkeo


def get_asm(emg, window_len, step_len, fs):
    window_sample = int(window_len * fs)
    step_sample = int(step_len * fs)
    (Nsample, Nchannel) = emg.shape

    asm = np.zeros((6, Nchannel))
    fea_idx = -1
    for i in range(0, Nsample - window_sample + 1, step_sample):
        fea_idx += 1
        for j in range(Nchannel):
            emg_window = emg[i:i + window_sample, j]
            asm[fea_idx, j] = my_asm(emg_window)
    return asm


def feature_train_normalize(feature_train, pca_active, dim):
    feature_train_concat = feature_train
    mean_val = np.mean(feature_train_concat, axis=1)
    std_val = np.std(feature_train_concat, axis=1, ddof=1)

    if pca_active == 1:
        pass
    else:
        WPCA = 0

    feature_train_norm = np.zeros(feature_train_concat.shape)
    for i in range(feature_train_concat.shape[1]):
        feature_train_norm[:, i] = (feature_train_concat[:, i] - mean_val) / std_val

    return feature_train_norm, mean_val, std_val, WPCA


def online_train(path, subject, num_train, snr_channel_mean):
    num_gesture = 5
    num_action = num_gesture * num_train
    filename = path + subject + '_label/' + subject + '_label.mat'
    file = scio.loadmat(file_name=filename)
    randIndex = file['randIndex']
    idx_train_tmp = [1, 2, 3]

    gesture_len = 1
    window_len = 0.5
    step_len = 0.1
    num_prd = int((gesture_len - window_len) / step_len + 1)
    pca_active = 0
    dim = 20
    fs_emg = 1000
    Nsample = fs_emg * gesture_len
    snr_thresh = 10

    # label_dyn = [0] * 25
    # randIndex=list(randIndex[0])
    # label_temp=[0]*6*num_gesture
    # for i in range(num_gesture):
    #     label_temp[6*i:6*i+6]=[randIndex[i]]*6
    #
    # label_train_tmp_1 = [1] * 6 * num_action + [0] * 6 * num_action
    # label_train_tmp_2 = label_temp*num_train

    filename = path + subject + '/' + subject + '_emg_rest.mat'
    file = scio.loadmat(filename)
    emg_rest = file['emg_rest']
    filename = path + subject + '/' + subject + '_emg_dyn.mat'
    file = scio.loadmat(filename)
    emg_dyn = file['emg_dyn']

    randIndex = randIndex[0]
    randIndex = np.delete(randIndex,randIndex==5)
    label_train_tmp_1 = [1] * (num_action*num_prd) + [0] * (num_action*num_prd)
    label_train_tmp_2 = []
    for _ in range(num_train):
        label_train_tmp_2+=[randIndex[0]] * 6 + [randIndex[1]] * 6 + [randIndex[2]] * 6 + [randIndex[3]] * 6 + [randIndex[4]] * 6
    # label_train_tmp_2 = [randIndex[0]] * 6 + [randIndex[1]] * 6 + [randIndex[2]] * 6 + [randIndex[3]] * 6 + \
    #                     [randIndex[4]] * 6 + [randIndex[0]] * 6 + [randIndex[1]] * 6 + [randIndex[2]] * 6 + \
    #                     [randIndex[3]] * 6 + [randIndex[4]] * 6 + [randIndex[0]] * 6 + [randIndex[1]] * 6 + [
    #                         randIndex[2]] * 6 + \
    #                     [randIndex[3]] * 6 + [randIndex[4]] * 6 + [randIndex[0]] * 6 + [randIndex[1]] * 6 + [
    #                         randIndex[2]] * 6 + \
    #                     [randIndex[3]] * 6 + [randIndex[4]] * 6 + [randIndex[0]] * 6 + [randIndex[1]] * 6 + [
    #                         randIndex[2]] * 6 + [randIndex[3]] * 6 + [randIndex[4]] * 6

    # idx_rel_train_tmp = []
    # for i in range(5):
    #     idx_rel_train_tmp.append(label_dyn.index(i))
    # idx_train = []
    # for i in range(5):
    #     idx_train.append(idx_rel_train_tmp[i])
    #     idx_train.append(idx_rel_train_tmp[i] + 1)
    #     idx_train.append(idx_rel_train_tmp[i] + 2)
    #     idx_train.append(idx_rel_train_tmp[i] + 3)
    #     idx_train.append(idx_rel_train_tmp[i] + 4)

    # label_train_1 = [1] * len(idx_train) + [0] * len(idx_train)
    # label_train_2 = [1]*3 + [2]*3 + [3]*3 + [4]*3 + [5]*3

    emg_train_rel = emg_dyn
    emg_train_irr = emg_rest

    idx_snr = snr_channel_mean > snr_thresh
    num_channel_snr = sum(idx_snr == 1)

    EMGfeature_train_rel = np.zeros((4*num_channel_snr, 6*num_action))
    for j in range(num_train*num_gesture):
        emg = emg_train_rel[j]
        emg = emg[:, idx_snr]
        var = get_var(emg, window_len, step_len, fs_emg)
        vo = get_vo(emg, window_len, step_len, fs_emg)
        ltkeo = get_ltkeo(emg, window_len, step_len, fs_emg)
        asm = get_asm(emg, window_len, step_len, fs_emg)
        for k in range(6):
            EMGfeature_train_rel[:, j * num_prd + k] = np.concatenate((var[k, :], vo[k, :], ltkeo[k, :], asm[k, :]),
                                                                      axis=0)

    EMGfeature_train_irr = np.zeros((4*num_channel_snr, 6*num_action))
    for j in range(num_train*num_gesture):
        emg = emg_train_irr[j][-1 - 1000 + 1:]
        emg = emg[:, idx_snr]
        var = get_var(emg, window_len, step_len, fs_emg)
        vo = get_vo(emg, window_len, step_len, fs_emg)
        ltkeo = get_ltkeo(emg, window_len, step_len, fs_emg)
        asm = get_asm(emg, window_len, step_len, fs_emg)
        for k in range(6):
            EMGfeature_train_irr[:, j * num_prd + k] = np.concatenate((var[k, :], vo[k, :], ltkeo[k, :], asm[k, :]),
                                                                      axis=0)

    EMGfeature_train_1 = np.concatenate((EMGfeature_train_rel, EMGfeature_train_irr), axis=1)
    # 归一化待补充
    EMGfeature_train_norm_1, EMGmean_val_1, EMGstd_val_1, EMGWPCA_1 = feature_train_normalize(EMGfeature_train_1,
                                                                                              pca_active, dim)
    EMGfeature_train_norm_2, EMGmean_val_2, EMGstd_val_2, EMGWPCA_2 = feature_train_normalize(EMGfeature_train_rel,
                                                                                              pca_active, dim)

    EMGmdl_1 = svm_train(label_train_tmp_1, EMGfeature_train_norm_1.T, '-t 0')
    EMGmdl_2 = svm_train(label_train_tmp_2, EMGfeature_train_norm_2.T, '-t 0')

    svm_save_model(subject + 'mdl1', EMGmdl_1)
    svm_save_model(subject + 'mdl2', EMGmdl_2)

    data = {'EMGmean_val_1': EMGmean_val_1, 'EMGstd_val_1': EMGstd_val_1, 'EMGWPCA_1': EMGWPCA_1,
            'EMGmean_val_2': EMGmean_val_2, 'EMGstd_val_2': EMGstd_val_2, 'EMGWPCA_2': EMGWPCA_2,
            'pca_active': pca_active, 'dim': dim, 'idx_snr': idx_snr}
    filename = path + subject + '/' + subject + '_train_data.mat'
    scio.savemat(filename, data)


def online_test(subject, path, EMGdata_test, Grodata_test, Accdata_test, label2_last, cd_last, cd_time, time_now, mode):
    mdl1 = svm_load_model (subject + 'mdl1')
    mdl2 = svm_load_model (subject + 'mdl2')
    file = scio.loadmat(file_name=path)

    gesture_len = 1
    window_len = 0.5
    step_len = 0.1
    num_prd = int((gesture_len - window_len) / step_len + 1)
    fs_emg = 1000
    Nsample = fs_emg * gesture_len
    Nsample_step = int(fs_emg * step_len)
    label1 = []
    label2 = []

    num_channel_snr = sum(file['idx_snr'][0] == 1)
    EMGfeature_test = np.zeros((4*num_channel_snr, 6))
    emg = EMGdata_test
    emg = emg[:, file['idx_snr'][0]==1]
    # np.savetxt('./emg.txt',np.shape(emg))
    # np.savetxt('./emg.txt',file['idx_snr'][0])
    np.savetxt('./emg.txt',emg)
    var = get_var(emg, window_len, step_len, fs_emg)
    np.savetxt('./var.txt',var)
    vo = get_vo(emg, window_len, step_len, fs_emg)
    ltkeo = get_ltkeo(emg, window_len, step_len, fs_emg)
    asm = get_asm(emg, window_len, step_len, fs_emg)
    for k in range(6):
        EMGfeature_test[:, k] = np.concatenate((var[k, :], vo[k, :], ltkeo[k, :], asm[k, :]), axis=0)

    # 归一化待补充
    EMGfeature_test_norm_1 = feature_test_normalize(EMGfeature_test, file['EMGmean_val_1'], file['EMGstd_val_1'],
                                                    file['pca_active'], file['dim'], file['EMGWPCA_1'])

    label_zero = np.zeros((6, 1))
    EMGlabel_predict_tmp_1, _, _ = svm_predict(label_zero, EMGfeature_test_norm_1.T, mdl1)

    label_predict_tmp_1 = EMGlabel_predict_tmp_1
    label_predict_1 = max(label_predict_tmp_1, key=label_predict_tmp_1.count)

    if label_predict_1 == 1:
        # 归一化待补充
        EMGfeature_test_norm_2 = feature_test_normalize(EMGfeature_test, file['EMGmean_val_2'], file['EMGstd_val_2'],
                                                    file['pca_active'], file['dim'], file['EMGWPCA_2'])

        EMGlabel_predict_tmp_2, _, _ = svm_predict(label_zero, EMGfeature_test_norm_2.T, mdl2)

        label_predict_tmp_2 = EMGlabel_predict_tmp_2
        for i in range(len(label_predict_tmp_2)):
            label_predict_tmp_2[i] += 1
        ges_mode_2 = max(label_predict_tmp_2, key=label_predict_tmp_2.count)
        ges_times_2 = label_predict_tmp_2.count(ges_mode_2)

        # if ges_times_2 == 6:
        #     label_predict_2 = ges_mode_2
        # else:
        #     label_predict_2 = 0

        if mode == 0:   # high-accuracy
            if ges_times_2 == 6:
                label_predict_2 = ges_mode_2
            else:
                label_predict_2 = 0
        elif mode == 1:  # high-sensitivity
            if ges_times_2 >= 5:
                label_predict_2 = ges_mode_2
            else:
                label_predict_2 = 0
    else:
        label_predict_2 = 0

    label2_now = label_predict_2

    if cd_last == 1:
        label_now = 0
        if time_now < 1500:
            time_now = time_now + 60000
        if time_now - cd_time >= 1500:
            cd_new = 0
        else:
            cd_new = 1
    if cd_last == 0:
        if label2_last * label2_now > 0 and label2_last == label2_now:
            label_now = label_predict_2
            cd_new = 1
            cd_time = time_now
        else:
            label_now = 0
            cd_new = 0

    return label_now, label2_now, cd_new, cd_time


def feature_test_normalize(feature_test, mean_train, std_train, pca_active, dim, WPCA):
    feature_test_norm = np.zeros(feature_test.shape)
    for i in range(feature_test.shape[1]):
        feature_test_norm[:, i] = (feature_test[:, i] - mean_train) / std_train

    return feature_test_norm


def deep_forest_online_test(subject, path, EMGdata_test, Grodata_test, Accdata_test, label2_last, cd_last, cd_time, time_now, mode, forest_mdl_test,mdl1):
    # mdl1 = svm_load_model(subject + 'mdl1')
    # mdl2 = svm_load_model(subject + 'mdl2')
    # save_path = path + subject + '/' + subject + 'forest'
    # forest_mdl_test = CascadeForestClassifier(max_layers=5, n_estimators=2, n_trees=100, random_state=1, n_jobs=-1,
    #                                            backend="sklearn", use_predictor=True)
    # forest_mdl_test.load(save_path)
    # file = scio.loadmat(file_name=path)
    feature_st = time.time()
    gesture_len = 1
    window_len = 0.5
    step_len = 0.1
    num_prd = int((gesture_len - window_len) / step_len + 1)
    fs_emg = 1000
    Nsample = fs_emg * gesture_len
    Nsample_step = int(fs_emg * step_len)
    label1 = []
    label2 = []

    # num_channel_snr = sum(file['idx_snr'][0] == 1)
    EMGfeature_test = np.zeros((4*8, 6))
    emg = EMGdata_test
    # emg = emg[:, file['idx_snr'][0]]
    var = get_var(emg, window_len, step_len, fs_emg)
    vo = get_vo(emg, window_len, step_len, fs_emg)
    ltkeo = get_ltkeo(emg, window_len, step_len, fs_emg)
    asm = get_asm(emg, window_len, step_len, fs_emg)
    for k in range(6):
        EMGfeature_test[:, k] = np.concatenate((var[k, :], vo[k, :], ltkeo[k, :], asm[k, :]), axis=0)

    # 归一化待补充
    # EMGfeature_test_norm_1 = feature_test_normalize(EMGfeature_test, file['EMGmean_val_1'], file['EMGstd_val_1'],
    #                                                 file['pca_active'], file['dim'], file['EMGWPCA_1'])
    print('feature time: ',time.time()-feature_st)
    label_zero = np.zeros((6, 1))
    svm_st = time.time()
    EMGlabel_predict_tmp_1, _, _ = svm_predict(label_zero, EMGfeature_test.T, mdl1)
    print('svm time: ', time.time() - svm_st)

    label_predict_tmp_1 = EMGlabel_predict_tmp_1
    label_predict_1 = max(label_predict_tmp_1, key=label_predict_tmp_1.count)

    if label_predict_1 == 1 and label_predict_tmp_1.count(1) >= 5:
        # 归一化待补充
        # EMGfeature_test_norm_2 = feature_test_normalize(EMGfeature_test, file['EMGmean_val_2'], file['EMGstd_val_2'],
        #                                             file['pca_active'], file['dim'], file['EMGWPCA_2'])

        # EMGlabel_predict_tmp_2, _, _ = svm_predict(label_zero, EMGfeature_test_norm_2.T, mdl2)
        forest_st = time.time()
        EMGlabel_predict_tmp_2 = forest_mdl_test.predict(EMGfeature_test.T)

        ges_mode_2 = np.argmax(np.bincount(EMGlabel_predict_tmp_2.astype(int)))
        ges_times_2 = np.max(np.bincount(EMGlabel_predict_tmp_2.astype(int)))
        label_predict_2 = ges_mode_2 + 1
        print('forest time: ', time.time() - forest_st)
        # label_predict_tmp_2 = EMGlabel_predict_tmp_2
        # for i in range(len(label_predict_tmp_2)):
        #     label_predict_tmp_2[i] += 1
        # ges_mode_2 = max(label_predict_tmp_2, key=label_predict_tmp_2.count)
        # ges_times_2 = label_predict_tmp_2.count(ges_mode_2)

        # if ges_times_2 == 6:
        #     label_predict_2 = ges_mode_2
        # else:
        #     label_predict_2 = 0

        # if mode == 0:   # high-accuracy
        #     if ges_times_2 == 6:
        #         label_predict_2 = ges_mode_2
        #     else:
        #         label_predict_2 = 0
        # elif mode == 1:  # high-sensitivity
        #     if ges_times_2 >= 5:
        #         label_predict_2 = ges_mode_2
        #     else:
        #         label_predict_2 = 0
    else:
        label_predict_2 = 0

    label2_now = label_predict_2
    inte_st = time.time()
    if cd_last == 1:
        label_now = 0
        if time_now < 1500:
            time_now = time_now + 60000
        if time_now - cd_time >= 1500:
            cd_new = 0
        else:
            cd_new = 1
    if cd_last == 0:
        if label2_last * label2_now > 0 and label2_last == label2_now:
            label_now = label_predict_2
            cd_new = 1
            cd_time = time_now
        else:
            label_now = 0
            cd_new = 0
    print('cd predict time: ', time.time()-inte_st)

    return label_now, label2_now, cd_new, cd_time

